from __future__ import annotations

import io
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def normalize_header(value: str) -> str:
    if not value:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace("à", "a")
        .replace("è", "e")
        .replace("é", "e")
        .replace("ì", "i")
        .replace("ò", "o")
        .replace("ù", "u")
    )


def safe_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    s = str(value).strip()
    if not s:
        return default
    s = s.replace(" ", "").replace("€", "")
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return default


def safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    s = str(value).strip()
    if not s:
        return default
    s_upper = s.upper()
    if s_upper in {"N", "NA", "N/A", "ND", "-", "NON DISPONIBILE"}:
        return default
    s_clean = s.replace(" ", "")
    if s_clean.replace(".", "").replace(",", "").isdigit():
        s_clean = s_clean.replace(".", "").replace(",", "")
    try:
        return int(s_clean)
    except ValueError:
        return default


def read_xlsx_table(file_bytes: bytes) -> tuple[dict[str, int], list[tuple[Any, ...]]]:
    if not file_bytes:
        raise ValueError("empty file")

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except InvalidFileException as exc:
        raise ValueError("invalid excel") from exc

    if not wb.sheetnames:
        raise ValueError("no sheets")

    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map: dict[str, int] = {}
    for idx, header in enumerate(header_row or []):
        key = normalize_header(header)
        if not key:
            continue
        header_map[key] = idx

    rows: list[tuple[Any, ...]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if not any(row):
            continue
        rows.append(tuple(row))

    return header_map, rows


def get_cell(
    row: Iterable[Any],
    header_map: dict[str, int],
    *candidates: str,
    default: Any = None,
) -> Any:
    row_list = list(row)
    for cand in candidates:
        cand_norm = normalize_header(cand)
        idx = header_map.get(cand_norm)
        if idx is not None and idx < len(row_list):
            return row_list[idx]
    return default
